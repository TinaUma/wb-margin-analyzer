from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.models.models import Product

_FILLS = {
    "green": PatternFill("solid", fgColor="C6EFCE"),
    "yellow": PatternFill("solid", fgColor="FFEB9C"),
    "red": PatternFill("solid", fgColor="FFC7CE"),
}

_HEADERS = [
    "Артикул",
    "Название",
    "Категория",
    "Цена продажи",
    "Себестоимость",
    "Вес, кг",
    "Продано",
    "Возвраты",
    "Выручка",
    "Комиссия",
    "Логистика",
    "Затр. закупки",
    "Прибыль",
    "Маржа %",
    "Зона",
]

_BOLD = Font(bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")


def build_report_xlsx(
    products: list[Product],
    ai_interpretation: str | None,
) -> bytes:
    wb = Workbook()

    # ── Sheet 1: продукты ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Маржинальность"

    ws1.append(_HEADERS)
    for cell in ws1[1]:
        cell.font = _BOLD

    for p in products:
        ws1.append(
            [
                p.article,
                p.name,
                p.category,
                float(p.sale_price),
                float(p.purchase_price),
                float(p.weight),
                p.sold,
                p.returns,
                float(p.net_revenue),
                float(p.commission),
                float(p.logistics_total),
                float(p.purchase_cost),
                float(p.profit),
                float(p.margin_pct),
                p.margin_zone,
            ]
        )
        fill = _FILLS.get(p.margin_zone, _FILLS["yellow"])
        for cell in ws1[ws1.max_row]:
            cell.fill = fill

    # Approximate column widths
    col_widths = [12, 30, 18, 14, 14, 10, 10, 10, 14, 12, 12, 14, 12, 10, 10]
    for i, width in enumerate(col_widths, start=1):
        ws1.column_dimensions[ws1.cell(1, i).column_letter].width = width

    # ── Sheet 2: AI интерпретация ─────────────────────────────────────────────
    ws2 = wb.create_sheet("AI Интерпретация")

    ws2.column_dimensions["A"].width = 90

    ws2["A1"] = "AI-анализ маржинальности"
    ws2["A1"].font = _BOLD

    ws2["A2"] = (
        ai_interpretation
        or "Интерпретация не была сгенерирована. Откройте анализ и нажмите «Получить AI-интерпретацию»."
    )
    ws2["A2"].alignment = _WRAP
    ws2.row_dimensions[2].height = 300

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
