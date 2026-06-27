"""Tests for GET /analyses/{id}/export — Task 9 AC."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from decimal import Decimal
from unittest.mock import MagicMock, AsyncMock

import openpyxl
import pytest
from fastapi.testclient import TestClient

from backend.api.v1.deps import get_current_user
from backend.db import get_db
from backend.main import app
from backend.models.models import AnalysisSession, Product, Report
from tests.conftest import FAKE_USER

_NOW = datetime(2026, 6, 24, 12, 0, 0, tzinfo=timezone.utc)
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Fake model builders ───────────────────────────────────────────────────────


def _fake_session(user_id: int = 1) -> AnalysisSession:
    s = AnalysisSession()
    s.id = 42
    s.user_id = user_id
    s.status = "done"
    s.created_at = _NOW
    s.files_meta = {}
    return s


def _fake_product(zone: str = "green") -> Product:
    p = Product()
    p.id = 1
    p.session_id = 42
    p.article = "SKU-001"
    p.name = "Тестовый товар"
    p.category = "Прочее"
    p.purchase_price = Decimal("500.00")
    p.sale_price = Decimal("1000.00")
    p.weight = Decimal("0.50")
    p.sold = 10
    p.returns = 0
    p.net_revenue = Decimal("10000.00")
    p.commission = Decimal("1800.00")
    p.logistics_total = Decimal("400.00")
    p.purchase_cost = Decimal("5000.00")
    p.profit = Decimal("2800.00")
    p.margin_pct = Decimal("28.00")
    p.margin_zone = zone
    return p


def _fake_report(ai_text: str | None = "## Диагноз\nОК") -> Report:
    r = Report()
    r.id = 1
    r.session_id = 42
    r.ai_interpretation = ai_text
    r.chat_history = []
    r.created_at = _NOW
    return r


def _make_db(session_obj=None, products=None, report_obj=None):
    db = AsyncMock()
    db.add = MagicMock()
    db.get.return_value = session_obj

    products_result = MagicMock()
    products_result.scalars.return_value = products or []

    report_result = MagicMock()
    report_result.scalar_one_or_none.return_value = report_obj

    db.execute.side_effect = [products_result, report_result]
    return db


def _db_dep(db):
    async def _override():
        yield db

    return _override


# ── AC-1: returns valid .xlsx ─────────────────────────────────────────────────


def test_export_returns_valid_xlsx():
    db = _make_db(
        session_obj=_fake_session(),
        products=[_fake_product("green"), _fake_product("red")],
        report_obj=_fake_report(),
    )
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/42/export")
    finally:
        app.dependency_overrides.clear()

    assert resp.status_code == 200
    assert resp.headers["content-type"] == _XLSX
    # Verify it's a parseable workbook
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb is not None


# ── AC-2: Sheet 1 exists with correct columns ─────────────────────────────────


def test_export_sheet1_has_products():
    db = _make_db(
        session_obj=_fake_session(),
        products=[_fake_product("green")],
        report_obj=_fake_report(),
    )
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/42/export")
    finally:
        app.dependency_overrides.clear()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Маржинальность" in wb.sheetnames
    ws = wb["Маржинальность"]
    # Header row
    headers = [ws.cell(1, c).value for c in range(1, 16)]
    assert "Артикул" in headers
    assert "Маржа %" in headers
    # Data row has the article
    assert ws.cell(2, 1).value == "SKU-001"


# ── AC-2: rows are colored by margin_zone ────────────────────────────────────


def test_export_rows_are_colored():
    products = [_fake_product("green"), _fake_product("yellow"), _fake_product("red")]
    products[1].article = "SKU-002"
    products[2].article = "SKU-003"

    db = _make_db(session_obj=_fake_session(), products=products, report_obj=None)
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/42/export")
    finally:
        app.dependency_overrides.clear()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Маржинальность"]
    green_fill = ws.cell(2, 1).fill.fgColor.rgb
    yellow_fill = ws.cell(3, 1).fill.fgColor.rgb
    red_fill = ws.cell(4, 1).fill.fgColor.rgb

    assert green_fill.endswith("C6EFCE"), f"Expected green, got {green_fill}"
    assert yellow_fill.endswith("FFEB9C"), f"Expected yellow, got {yellow_fill}"
    assert red_fill.endswith("FFC7CE"), f"Expected red, got {red_fill}"


# ── AC-3: Sheet 2 exists with AI text ────────────────────────────────────────


def test_export_sheet2_has_ai_interpretation():
    ai_text = "## Диагноз\nВсё хорошо"
    db = _make_db(
        session_obj=_fake_session(),
        products=[_fake_product()],
        report_obj=_fake_report(ai_text=ai_text),
    )
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/42/export")
    finally:
        app.dependency_overrides.clear()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "AI Интерпретация" in wb.sheetnames
    ws = wb["AI Интерпретация"]
    assert ws["A2"].value == ai_text


# ── AC-4: filename contains analysis date ─────────────────────────────────────


def test_export_filename_contains_date():
    db = _make_db(session_obj=_fake_session(), products=[], report_obj=None)
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/42/export")
    finally:
        app.dependency_overrides.clear()

    disposition = resp.headers.get("content-disposition", "")
    assert "report_2026-06-24.xlsx" in disposition


# ── AC-5 (negative): 404 for missing analysis ────────────────────────────────


def test_export_404_for_missing_analysis():
    db = _make_db(session_obj=None)
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/999/export")
    finally:
        app.dependency_overrides.clear()

    assert resp.status_code == 404


# ── AC-6 (negative): 403 for another user's analysis ─────────────────────────


def test_export_403_for_other_user():
    db = _make_db(session_obj=_fake_session(user_id=999))
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER  # id=1
    app.dependency_overrides[get_db] = _db_dep(db)
    try:
        with TestClient(app) as c:
            resp = c.get("/api/v1/analyses/42/export")
    finally:
        app.dependency_overrides.clear()

    assert resp.status_code == 403
